import io
import json
import threading
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from app.auth.clerk_auth import ClerkAuthUser, get_clerk_user

_STORAGE = Path(__file__).resolve().parent.parent.parent / "storage" / "solicitacoes.json"
_lock = threading.Lock()

router = APIRouter(prefix="/dashboard", tags=["dashboard"])


def _read() -> list:
    with _lock:
        if not _STORAGE.exists():
            return []
        try:
            return json.loads(_STORAGE.read_text(encoding="utf-8")) or []
        except (json.JSONDecodeError, OSError):
            return []


def _aprovadas(items: list) -> list:
    return [s for s in items if s.get("status") == "aprovado"]


def _fmt_mes(iso: str) -> str:
    try:
        return datetime.fromisoformat(iso).strftime("%Y-%m")
    except Exception:
        return "desconhecido"


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/resumo")
def resumo(_: ClerkAuthUser = Depends(get_clerk_user)):
    items = _read()
    aprov = _aprovadas(items)

    desconto_medio = (
        sum(s.get("desconto_percentual", 0) for s in aprov) / len(aprov) if aprov else 0.0
    )
    economia_total = sum(
        (s.get("diaria_atual", 0) - s.get("preco_sugerido", 0)) for s in aprov
    )

    contagem: dict[str, dict] = {}
    for s in aprov:
        lid = s.get("listing_id", "")
        if lid not in contagem:
            contagem[lid] = {
                "listing_id": lid,
                "listing_nome": s.get("listing_nome", lid),
                "total_aprovadas": 0,
                "soma_desconto": 0.0,
            }
        contagem[lid]["total_aprovadas"] += 1
        contagem[lid]["soma_desconto"] += s.get("desconto_percentual", 0)

    top = sorted(contagem.values(), key=lambda x: x["total_aprovadas"], reverse=True)[:5]
    top_imoveis = [
        {
            "listing_id": t["listing_id"],
            "listing_nome": t["listing_nome"],
            "total_aprovadas": t["total_aprovadas"],
            "desconto_medio": round(t["soma_desconto"] / t["total_aprovadas"], 2) if t["total_aprovadas"] else 0.0,
        }
        for t in top
    ]

    return {
        "total_solicitacoes": len(items),
        "aprovadas": len(aprov),
        "rejeitadas": sum(1 for s in items if s.get("status") == "rejeitado"),
        "pendentes": sum(1 for s in items if s.get("status") == "pendente"),
        "desconto_medio_percentual": round(desconto_medio, 2),
        "economia_total": round(economia_total, 2),
        "top_imoveis": top_imoveis,
    }


@router.get("/historico")
def historico(_: ClerkAuthUser = Depends(get_clerk_user)):
    aprov = _aprovadas(_read())

    por_mes: dict[str, dict] = defaultdict(lambda: {"total": 0, "soma_desconto": 0.0, "soma_economia": 0.0})
    for s in aprov:
        mes = _fmt_mes(s.get("criado_em", ""))
        por_mes[mes]["total"] += 1
        por_mes[mes]["soma_desconto"] += s.get("desconto_percentual", 0)
        por_mes[mes]["soma_economia"] += s.get("diaria_atual", 0) - s.get("preco_sugerido", 0)

    return sorted(
        [
            {
                "mes": mes,
                "total": d["total"],
                "desconto_medio": round(d["soma_desconto"] / d["total"], 2) if d["total"] else 0.0,
                "economia": round(d["soma_economia"], 2),
            }
            for mes, d in por_mes.items()
        ],
        key=lambda x: x["mes"],
    )


@router.get("/score/{listing_id}")
def score(listing_id: str, _: ClerkAuthUser = Depends(get_clerk_user)):
    items = _read()
    imovel = [s for s in items if s.get("listing_id") == listing_id]
    aprov = [s for s in imovel if s.get("status") == "aprovado"]

    nome = next((s.get("listing_nome", listing_id) for s in imovel), listing_id)
    total = len(imovel)
    n_aprov = len(aprov)
    desc_medio = (
        sum(s.get("desconto_percentual", 0) for s in aprov) / n_aprov if n_aprov else 0.0
    )

    freq_score = min(total / 10, 1.0) * 30          # até 30 pts
    taxa_score = (n_aprov / total if total else 0) * 40  # até 40 pts
    desc_score = max(0, 1 - desc_medio / 30) * 30   # até 30 pts (menor desconto = melhor)
    score_val = round(freq_score + taxa_score + desc_score)

    return {
        "listing_id": listing_id,
        "listing_nome": nome,
        "score": score_val,
        "total": total,
        "aprovadas": n_aprov,
        "desconto_medio": round(desc_medio, 2),
    }


@router.get("/export/excel")
def export_excel(_: ClerkAuthUser = Depends(get_clerk_user)):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    aprov = _aprovadas(_read())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aprovações"

    header_fill = PatternFill("solid", fgColor="C0392B")
    header_font = Font(bold=True, color="FFFFFF")
    cols = ["Data", "Imóvel", "Atendente", "Diária Atual (R$)", "Desconto (%)", "Preço Sugerido (R$)", "Repasse (R$)", "Aprovador"]
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in aprov:
        data = ""
        try:
            data = datetime.fromisoformat(s.get("criado_em", "")).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        ws.append([
            data,
            s.get("listing_nome", s.get("listing_id", "")),
            s.get("solicitante_nome") or s.get("solicitante_email", ""),
            s.get("diaria_atual", 0),
            s.get("desconto_percentual", 0),
            s.get("preco_sugerido", 0),
            s.get("repasse_resultante", 0),
            s.get("aprovador_nome") or s.get("aprovador_email", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=aprovacoes.xlsx"},
    )


@router.get("/export/pdf")
def export_pdf(_: ClerkAuthUser = Depends(get_clerk_user)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    items = _read()
    aprov = _aprovadas(items)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = []

    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#C0392B")
    story.append(Paragraph("WeCare — Relatório de Aprovações", title_style))
    story.append(Spacer(1, 6 * mm))

    desc_medio = (
        sum(s.get("desconto_percentual", 0) for s in aprov) / len(aprov) if aprov else 0.0
    )
    economia = sum((s.get("diaria_atual", 0) - s.get("preco_sugerido", 0)) for s in aprov)

    resumo_data = [
        ["Total de solicitações", str(len(items))],
        ["Aprovadas", str(len(aprov))],
        ["Rejeitadas", str(sum(1 for s in items if s.get("status") == "rejeitado"))],
        ["Desconto médio", f"{desc_medio:.2f}%"],
        ["Economia total", f"R$ {economia:,.2f}"],
    ]
    resumo_tbl = Table(resumo_data, colWidths=[60 * mm, 40 * mm])
    resumo_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F5F5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#FDF2F2")]),
    ]))
    story.append(resumo_tbl)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Solicitações aprovadas", styles["Heading2"]))
    story.append(Spacer(1, 3 * mm))

    header = ["Data", "Imóvel", "Atendente", "Diária (R$)", "Desc.%", "Preço sug. (R$)", "Repasse (R$)", "Aprovador"]
    rows = [header]
    for s in aprov:
        data = ""
        try:
            data = datetime.fromisoformat(s.get("criado_em", "")).strftime("%d/%m/%Y")
        except Exception:
            pass
        rows.append([
            data,
            (s.get("listing_nome") or s.get("listing_id", ""))[:25],
            (s.get("solicitante_nome") or s.get("solicitante_email", ""))[:20],
            f"{s.get('diaria_atual', 0):.2f}",
            f"{s.get('desconto_percentual', 0):.1f}%",
            f"{s.get('preco_sugerido', 0):.2f}",
            f"{s.get('repasse_resultante', 0):.2f}",
            (s.get("aprovador_nome") or s.get("aprovador_email", ""))[:20],
        ])

    col_widths = [25 * mm, 50 * mm, 38 * mm, 22 * mm, 16 * mm, 28 * mm, 24 * mm, 38 * mm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C0392B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDF2F2")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=aprovacoes.pdf"},
    )
